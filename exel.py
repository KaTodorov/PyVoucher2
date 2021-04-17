
import openpyxl
import xlsxwriter
from datetime import date
import random
import math
import os
from base64 import b64encode

# parameters of 100 vouchers
times_before = 0  # first voucher number
keys = ["№", "Promotion", "Validity", "Code"]
verification = ["№", "Code", "Attendance"]
now = "31.03.2021"
promotion = "-50% OFF in the special hours between 10:00 - 18:00 in the working days "

# open working files

# raw format document
workbook1 = xlsxwriter.Workbook(
    str(times_before)+"-"+str(times_before+100)+'_raw.xlsx')
worksheet1 = workbook1.add_worksheet()

# verification document of the type "№"/"Code"/"Attendance" with 2 colmns
workbook2 = xlsxwriter.Workbook(
    str(times_before)+"-"+str(times_before+100)+'_print_verification.xlsx')
worksheet2 = workbook2.add_worksheet()

#
workbook = xlsxwriter.Workbook(
    str(times_before)+"-"+str(times_before+100)+"_"+'print.xlsx')
worksheet = workbook.add_worksheet()

# style of cells
borders_number = {"text_wrap": 1,
                  "border": 1,
                  "border_color": "#000000",
                  "font_size": 9

                  }
borders_date = {"border": 1,
                "border_color": "#000000",
                "num_format": "d.m.yyyy",
                "font_size": 9
                }


def add_format(given_workbook):
    cell_format = given_workbook.add_format(borders_number)
    cell_format.set_align('center')
    cell_format.set_align('vcenter')
    return cell_format


def add_date_format(given_workbook):
    cell_format = given_workbook.add_format(borders_date)
    cell_format.set_align('center')
    cell_format.set_align('vcenter')
    return cell_format


# add the formats to the opened documents
full_border = add_format(workbook1)
date_format = add_date_format(workbook1)
full_border_2 = add_format(workbook2)
date_format_2 = add_date_format(workbook2)
full_border_print = add_format(workbook)
date_print = add_format(workbook)

# code generation


def get_promo_code(num_chars=4):
    sec = os.urandom(6)
    token = b64encode(sec).decode('utf-8')
    return token[:6]


# put keys and format first raw
for colmn, data in enumerate(keys):
    worksheet1.write(0, colmn, data, full_border)

worksheet2.write(0, 0, verification[0], full_border_2)
worksheet2.write(0, 1, verification[1], full_border_2)
worksheet2.write(0, 2, verification[2], full_border_2)
worksheet2.write(0, 5, verification[0], full_border_2)
worksheet2.write(0, 6, verification[1], full_border_2)
worksheet2.write(0, 7, verification[2], full_border_2)


for row in range(1, 101):
    key = get_promo_code()
    worksheet1.write(row, 0, row + times_before, full_border)
    worksheet1.write(row, 1, promotion, full_border)
    worksheet1.write(row, 2, now, date_format)
    worksheet1.write(row, 3, key, full_border)
    if(row <= 50):
        worksheet2.write(row, 0, row+times_before, full_border_2)
        worksheet2.write(row, 1, key, full_border_2)
        worksheet2.write(row, 2, "Т", full_border_2)
    else:
        worksheet2.write(row-50, 5, row+times_before, full_border_2)
        worksheet2.write(row-50, 6, key, full_border_2)
        worksheet2.write(row-50, 7, "Т", full_border_2)

# set the size of each colmn
worksheet1.set_column(0, 0, 6)
worksheet1.set_column(1, 1, 20)
worksheet1.set_column(2, 2, 12)
worksheet1.set_column(3, 3, 8)


worksheet2.set_column(0, 0, 8)
worksheet2.set_column(1, 1, 12)
worksheet2.set_column(2, 2, 10)


worksheet2.set_column(5, 5, 8)
worksheet2.set_column(6, 6, 12)
worksheet2.set_column(7, 7, 10)


worksheet.set_column(0, 0, 6)
worksheet.set_column(1, 1, 20)
worksheet.set_column(2, 2, 12)
worksheet.set_column(3, 3, 8)
worksheet.set_column(4, 4, 1)
worksheet.set_column(5, 5, 6)
worksheet.set_column(6, 6, 20)
worksheet.set_column(7, 7, 12)
worksheet.set_column(8, 8, 8)
# close
workbook2.close()
workbook1.close()


# open the raw file to read
path = str(times_before)+"-"+str(times_before+100)+"_raw.xlsx"
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1, column=1)
max_row = sheet_obj.max_row
max_col = sheet_obj.max_column

second_colmn_row = 0
for k in range(2, max_row+1):
    if(k > 52):
        second_colmn_row = 2 + second_colmn_row
    for i in range(1, 5):
        cell_obj_title = sheet_obj.cell(row=1, column=i)
        cell_obj_row = sheet_obj.cell(row=k, column=i)
        if(k <= 51):
            worksheet.write(k*2-4, i-1, cell_obj_title.value,
                            full_border_print)
            if(i == 3):
                worksheet.write(k*2-3, i-1, cell_obj_row.value, date_print)
            else:
                worksheet.write(k*2-3, i-1, cell_obj_row.value,
                                full_border_print)
        else:
            worksheet.write(second_colmn_row, i+4,
                            cell_obj_title.value, full_border_print)
            if(i == 3):
                worksheet.write(second_colmn_row+1, i+4,
                                cell_obj_row.value, date_print)
            else:
                worksheet.write(second_colmn_row+1, i+4,
                                cell_obj_row.value, full_border_print)


# for l in range(1,)
#     for m in range(1,16):

workbook.close()
